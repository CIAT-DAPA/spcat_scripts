
from bs4 import BeautifulSoup
import openpyxl
from geo.Geoserver import Geoserver
import os
import sys
import pandas as pd

#geoserver data
dir_path = os.path.dirname(os.path.realpath(__file__))
parent_dir = os.path.abspath(os.path.join(dir_path, ".."))
path_database = os.path.abspath(os.path.join(parent_dir, 'database.xlsx'))

try:
    database = pd.read_excel(path_database)
except Exception as e:
    print('Problem reading database.xlsx file')
    print(e)
    sys.exit(1)

def get_parameter(name, df=database):
    value = df[df.parameter == name].iloc[0]['value']
    return str(value)

geoserver_url = get_parameter('geo_url')
username = get_parameter('user_gs')
password = get_parameter('pass_gs')
description=''
workspace = get_parameter('workspace_gs') #the workspace must exist on the geoserver
image_directory = get_parameter('path')
# create a geoserver instance
geo = Geoserver(geoserver_url, username=username, password=password)


# Create a new workbook
workbook = openpyxl.Workbook()

# Create a new sheet
worksheet = workbook.active
worksheet.title = 'Errors'

# write the headers ins columns
worksheet.cell(row=1, column=1, value='File')
worksheet.cell(row=1, column=2, value='Error')

# Inicialize the rows counter
row = 2

image_files = [f for f in os.listdir(image_directory) if os.path.isfile(os.path.join(image_directory, f)) and f.endswith('.tif')]

print('the process of import rasters has begun')
#initialize counters
success_count = 0
error_count = 0

for image_file in image_files:
    layer_name = os.path.splitext(image_file)[0]

    try:

        geo.create_coveragestore(layer_name=layer_name, path=os.path.join(image_directory, image_file), workspace=workspace)
        print(f'The {layer_name} file has been successfully loaded into the {workspace} workspace on your Geoserver.')
        success_count +=1
    except Exception as e:
        error_count += 1
        error=str(e)
        soup = BeautifulSoup(error, 'html.parser')
        p_tags = soup.find_all('p')
        # Extract error from the generated HTML tag
        for p in p_tags:
            if 'Description' in p.text:
                description = p.find('b').next_sibling.strip()
                break
        print(f'The importation of the {layer_name} file has failed')
        if description:
            print(description)
        else:
            ()
        worksheet.cell(row=row, column=1, value=image_file)
        worksheet.cell(row=row, column=2, value=str(description))
        row += 1
       
# save the workbook
workbook.save('errors_logs.xlsx')
#print results
print(f"{success_count } rasters have been uploaded to the Geoserver.")
print(f"Failed to upload {error_count} rasters.")

